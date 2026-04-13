from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter.ttk import Combobox
from tkinter import messagebox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib

# Create Excel file if not exists
file = pathlib.Path('MicroProject1_data.xlsx')

if not file.exists():
    file = Workbook()
    sheet = file.active

    headers = [
        "Name", "Date Of Birth", "Gender", "Address", "E-Mail",
        "Phone Number", "Dietary Preference", "Father's Name",
        "Mother's Name", "Guardian Name", "Parents' Address",
        "Parents' E-mail", "Parents' Phone Number",
        "Father's Occupation", "Mother's Occupation",
        "English Marks", "Physics Marks", "Chemistry Marks",
        "Maths Marks", "Optional Marks", "Aggregate Marks(PCM)",
        "Hobbies", "Achievement"
    ]

    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=header)

    file.save('MicroProject1_data.xlsx')


def submit():
    file = openpyxl.load_workbook('MicroProject1_data.xlsx')
    sheet = file.active

    row = sheet.max_row + 1

    name = f"{fnameVal.get()} {mnameVal.get()} {lnameVal.get()}"
    address = f"{loc.get()}, PIN={pin.get()}, {dist.get()}, {state.get()}, {country.get()}"
    p_address = f"{p_loc.get()}, PIN={p_pin.get()}, {p_dist.get()}, {p_state.get()}, {p_country.get()}"

    phy = p.get()
    chem = c.get()
    maths = mth.get()
    agg = (phy + chem + maths) / 3 if (phy and chem and maths) else 0

    data = [
        name, dob.get(), gender.get(), address, emAd.get(), phn.get(),
        diet.get(), father.get(), mother.get(), gaurd.get(),
        p_address, p_emAd.get(), p_phn.get(),
        foccu.get(), moccu.get(),
        e.get(), phy, chem, maths, op.get(),
        agg, hobbies.get(), achieve.get()
    ]

    for col, value in enumerate(data, start=1):
        sheet.cell(row=row, column=col, value=value)

    file.save('MicroProject1_data.xlsx')
    messagebox.showinfo("Success", "Data Saved Successfully!")


def clear():
    for var in [
        fnameVal, mnameVal, lnameVal, dob, pin, loc, dist, state, country,
        emAd, phn, father, mother, gaurd, p_pin, p_loc, p_dist,
        p_state, p_country, p_emAd, p_phn, foccu, moccu,
        hobbies, achieve
    ]:
        var.set("")

    for var in [p, c, mth, e, op]:
        var.set(0)


def on_configure(event):
    canvas.configure(scrollregion=canvas.bbox('all'))


# GUI Setup
window = Tk()
window.title("Admission Form")

canvas = Canvas(window)
canvas.pack(side=LEFT, fill=BOTH, expand=True)

vs = Scrollbar(window, orient="vertical", command=canvas.yview)
vs.pack(side=RIGHT, fill=Y)

hs = Scrollbar(window, orient="horizontal", command=canvas.xview)
hs.pack(side=BOTTOM, fill=X)

canvas.configure(yscrollcommand=vs.set, xscrollcommand=hs.set)

frame = Frame(canvas)
canvas.create_window((0, 0), window=frame, anchor="nw")

frame.bind("<Configure>", on_configure)

# Variables
fnameVal = StringVar()
mnameVal = StringVar()
lnameVal = StringVar()
dob = StringVar()
pin = StringVar()
loc = StringVar()
dist = StringVar()
state = StringVar()
country = StringVar()
emAd = StringVar()
phn = StringVar()
father = StringVar()
mother = StringVar()
gaurd = StringVar()

p_pin = StringVar()
p_loc = StringVar()
p_dist = StringVar()
p_state = StringVar()
p_country = StringVar()
p_emAd = StringVar()
p_phn = StringVar()

foccu = StringVar()
moccu = StringVar()

p = IntVar()
c = IntVar()
mth = IntVar()
e = IntVar()
op = IntVar()

hobbies = StringVar()
achieve = StringVar()

# UI (Shortened Example — You can expand similarly)
Label(frame, text="Admission Form", font="Arial 16 bold").grid(row=0, column=1)

Label(frame, text="First Name").grid(row=1, column=0)
Entry(frame, textvariable=fnameVal).grid(row=1, column=1)

Label(frame, text="Middle Name").grid(row=2, column=0)
Entry(frame, textvariable=mnameVal).grid(row=2, column=1)

Label(frame, text="Last Name").grid(row=3, column=0)
Entry(frame, textvariable=lnameVal).grid(row=3, column=1)

Label(frame, text="DOB").grid(row=4, column=0)
Entry(frame, textvariable=dob).grid(row=4, column=1)

Label(frame, text="Gender").grid(row=5, column=0)
gender = Combobox(frame, values=["Male", "Female"])
gender.grid(row=5, column=1)

Label(frame, text="Phone").grid(row=6, column=0)
Entry(frame, textvariable=phn).grid(row=6, column=1)

# Buttons
Button(frame, text="Submit", command=submit).grid(row=20, column=1)
Button(frame, text="Clear", command=clear).grid(row=20, column=2)
Button(frame, text="Exit", command=window.destroy).grid(row=20, column=3)

window.mainloop()
